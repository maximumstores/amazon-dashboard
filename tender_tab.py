"""
tender_tab.py — Streamlit таб для логістичного тендера

Інтеграція в dashboard.py:
  1. Скопіюй цей файл в корінь проекту (поряд з dashboard.py)
  2. В dashboard.py додай імпорт:
        from tender_tab import show_tender_tab
  3. В FBA Operations sidebar знайди функцію show_fba_operations() 
     і додай 6-й таб:
        tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
            "🚀 Shipments", "📋 Items", "🗑️ Removals", 
            "⚠️ Non-Compliance", "🏠 Inventory Health",
            "📋 Tender"   # <-- НОВИЙ
        ])
        ...
        with tab6:
            show_tender_tab()

Фічі:
  ✅ Фільтри (FC, status, date range, marketplace)
  ✅ Мультиселект shipments через checkbox column
  ✅ Live preview таблиці
  ✅ Агрегована статистика (total boxes/kg/CBM/units)
  ✅ Pick-up date input (ручний або календар)
  ✅ Download кнопка → Excel в форматі Олексія
  ✅ Завантаження квот перевізників (Модуль 2)
"""

import streamlit as st
import pandas as pd
import datetime
import io
import os
import re
import tempfile
import psycopg2
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ============================================
# DB helper
# ============================================
def _get_conn():
    url = os.getenv("DATABASE_URL") or st.secrets.get("DATABASE_URL", "")
    return psycopg2.connect(url)

@st.cache_data(ttl=300)
def fetch_tender_shipments():
    conn = _get_conn()
    try:
        df = pd.read_sql("""
            SELECT
                COALESCE(NULLIF(s.shipment_confirmation_id,''), s.shipment_id) AS fba_id,
                s.shipment_id,
                s.destination_fc                              AS fc,
                s.ship_to_line1                               AS line1,
                s.ship_to_city                                AS city,
                s.ship_to_state                               AS state,
                s.ship_to_postal                              AS postal,
                s.ship_to_country                             AS country,
                s.source_city                                 AS source_city,
                s.source_country                              AS source_country,
                s.delivery_window_start                       AS dw_start,
                s.delivery_window_end                         AS dw_end,
                s.status,
                s.name,
                s.marketplace,
                s.inbound_plan_id,
                s.placement_option_id,
                COUNT(DISTINCT b.box_id)                      AS box_count,
                COALESCE(SUM(NULLIF(b.weight_kg,'')::float),0)  AS total_kg,
                COALESCE(AVG(NULLIF(b.weight_kg,'')::float),0)  AS avg_kg,
                COALESCE(SUM(NULLIF(b.volume_cbm,'')::float),0) AS total_cbm,
                COALESCE(AVG(NULLIF(b.volume_cbm,'')::float),0) AS avg_cbm,
                COALESCE(SUM(NULLIF(bi.quantity,'')::int),0)    AS total_units
            FROM public.fba_inbound_shipments_v2 s
            LEFT JOIN public.fba_shipment_boxes b 
                ON b.inbound_plan_id = s.inbound_plan_id
            LEFT JOIN public.fba_shipment_box_items bi 
                ON bi.box_id = b.box_id 
                AND bi.inbound_plan_id = b.inbound_plan_id
            WHERE s.shipment_id != ''
            GROUP BY 
                s.shipment_confirmation_id, s.shipment_id, s.destination_fc,
                s.ship_to_line1, s.ship_to_city, s.ship_to_state, 
                s.ship_to_postal, s.ship_to_country,
                s.source_city, s.source_country,
                s.delivery_window_start, s.delivery_window_end,
                s.status, s.name, s.marketplace,
                s.inbound_plan_id, s.placement_option_id
            HAVING COUNT(DISTINCT b.box_id) > 0
            ORDER BY s.delivery_window_start NULLS LAST, s.destination_fc
        """, conn)
    finally:
        conn.close()
    return df

@st.cache_data(ttl=300)
def fetch_placement_fees():
    conn = _get_conn()
    try:
        df = pd.read_sql("""
            SELECT 
                placement_option_id,
                MAX(fee_amount)   AS fee_amount,
                MAX(fee_currency) AS fee_currency
            FROM public.fba_shipment_placement_fees
            WHERE fee_target = 'Placement Services'
              AND COALESCE(fee_amount, '') != ''
            GROUP BY placement_option_id
        """, conn)
    finally:
        conn.close()
    return df

# ============================================
# Excel builder (тендерний файл)
# ============================================
FONT_BOLD   = Font(name="Arial", size=11, bold=True)
FONT_HEADER = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FONT_NORMAL = Font(name="Arial", size=10)
FONT_SMALL  = Font(name="Arial", size=9, italic=True, color="666666")
FILL_HEADER    = PatternFill("solid", start_color="1F4E79")
FILL_SECTION   = PatternFill("solid", start_color="E7E6E6")
FILL_HIGHLIGHT = PatternFill("solid", start_color="FFF2CC")

def _format_dw(start, end):
    if not start or not end or pd.isna(start) or pd.isna(end):
        return ""
    try:
        s_str = str(start).replace("Z", "+00:00")
        e_str = str(end).replace("Z", "+00:00")
        s = datetime.datetime.fromisoformat(s_str)
        e = datetime.datetime.fromisoformat(e_str)
        if s.month == e.month and s.year == e.year:
            return f"{s.strftime('%b %d')} - {e.strftime('%d, %Y')}"
        return f"{s.strftime('%b %d')} - {e.strftime('%b %d, %Y')}"
    except Exception:
        return f"{str(start)[:10]} - {str(end)[:10]}"

def build_tender_excel(df_selected, pickup_date):
    wb = Workbook()
    ws = wb.active
    ws.title = "tender"
    widths = {"A": 18, "B": 12, "C": 8, "D": 6, "E": 6, "F": 10,
              "G": 6, "H": 6, "I": 6, "J": 6, "K": 28, "L": 32}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws["A1"] = "pick-up date"
    ws["A1"].font = FONT_BOLD
    ws["B1"] = pickup_date
    ws["B1"].font = FONT_BOLD
    ws["B1"].fill = FILL_HIGHLIGHT

    ws["A5"] = "Carton G.W. kg"
    ws["B5"] = "Carton N.W."
    ws["C5"] = "Carton volume"
    ws["D5"] = "Total G.W."
    ws["E5"] = "Total N.W."
    ws["F5"] = "Total volume, CBM"
    for col in "ABCDEF":
        ws[f"{col}5"].font = FONT_BOLD

    if len(df_selected) > 0:
        avg_kg  = df_selected["avg_kg"].mean()
        avg_cbm = df_selected["avg_cbm"].mean()
        total_kg  = df_selected["total_kg"].sum()
        total_cbm = df_selected["total_cbm"].sum()
        ws["A6"] = round(avg_kg, 2)
        ws["B6"] = round(avg_kg * 0.95, 2)
        ws["C6"] = round(avg_cbm, 4)
        ws["D6"] = round(total_kg, 2)
        ws["E6"] = round(total_kg * 0.95, 2)
        ws["F6"] = round(total_cbm, 4)
        for col in "ABCDEF":
            ws[f"{col}6"].fill = FILL_HIGHLIGHT

    ws["A9"] = "Amazon"
    ws["A9"].font = FONT_BOLD
    ws["A9"].fill = FILL_SECTION

    current_row = 11
    for _, ship in df_selected.iterrows():
        ws.cell(row=current_row, column=1, value=ship["fba_id"]).font = FONT_BOLD
        ws.cell(row=current_row, column=3, value=ship["fc"])
        ws.cell(row=current_row, column=7, value=int(ship["box_count"]))
        ws.cell(row=current_row, column=8, value="boxes")
        ws.cell(row=current_row, column=9, value=int(ship["total_units"]))
        ws.cell(row=current_row, column=10, value="units")

        fee_text = ""
        fee_amt  = ship.get("fee_amount", "") or ""
        fee_curr = ship.get("fee_currency", "") or "USD"
        if fee_amt and str(fee_amt) not in ("0", "0.0", ""):
            try:
                fee_text = f"Total placement fees: {fee_curr}${float(fee_amt):,.2f}"
            except (ValueError, TypeError):
                fee_text = f"Total placement fees: {fee_amt} {fee_curr}"
        if fee_text:
            ws.cell(row=current_row, column=11, value=fee_text).font = FONT_SMALL

        dw = _format_dw(ship.get("dw_start"), ship.get("dw_end"))
        if dw:
            ws.cell(row=current_row, column=12,
                    value=f"Delivery window: {dw}").font = FONT_SMALL

        line1  = ship.get("line1") or ""
        city   = (ship.get("city") or "").upper()
        state  = ship.get("state") or ""
        postal = ship.get("postal") or ""
        country = ship.get("country") or ""
        parts = [ship["fc"]]
        if line1:
            parts.append(f"{line1} {postal}".strip())
        parts.append(f"{city}, {state}")
        if country:
            parts.append("United States" if country == "US" else country)
        ship_to = "Ship to: " + " - ".join(parts)
        cell = ws.cell(row=current_row + 1, column=1, value=ship_to)
        cell.font = FONT_BOLD
        ws.merge_cells(start_row=current_row + 1, start_column=1,
                       end_row=current_row + 1, end_column=12)
        current_row += 3

    ws.freeze_panes = "A10"

    ws2 = wb.create_sheet("Delivery comparison")
    headers = ["FBA ID", "Destination FC", "Carrier", "Service type",
               "Cost (USD)", "Transit days", "Pick-up date", "Delivery date",
               "Container type", "Fuel surcharge", "Notes"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        ws2.column_dimensions[get_column_letter(col_idx)].width = 16
    for r_idx, ship in enumerate(df_selected.itertuples(), 2):
        ws2.cell(row=r_idx, column=1, value=ship.fba_id)
        ws2.cell(row=r_idx, column=2, value=ship.fc)
    ws2.freeze_panes = "A2"

    ws3 = wb.create_sheet("Transit Time Norms")
    ws3["A1"] = "Transit Time Norms (fill manually based on historical data)"
    ws3["A1"].font = FONT_BOLD
    ws3.merge_cells("A1:F1")
    norms_headers = ["Origin region", "Destination region", "Service type",
                     "Normal transit days", "Best case", "Worst case"]
    for col_idx, h in enumerate(norms_headers, 1):
        cell = ws3.cell(row=3, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        ws3.column_dimensions[get_column_letter(col_idx)].width = 20
    for r_idx, row in enumerate([
        ["China", "US East Coast", "Ocean LCL",      30, 25, 45],
        ["China", "US West Coast", "Ocean LCL",      20, 18, 35],
        ["China", "US East Coast", "Air Freight",     7,  5, 10],
        ["China", "US West Coast", "Air Freight",     5,  4,  8],
        ["US WH", "Amazon FC",     "Ground (LTL)",    3,  2,  7],
        ["US WH", "Amazon FC",     "Ground (Parcel)", 2,  1,  5],
    ], 4):
        for c_idx, val in enumerate(row, 1):
            ws3.cell(row=r_idx, column=c_idx, value=val)

    ws4 = wb.create_sheet("data for tender")
    data_headers = [
        "FBA ID", "Shipment ID", "FC", "Status",
        "City", "State", "Postal", "Country", "Address",
        "Source city", "Source country",
        "DW start", "DW end",
        "Boxes", "Total kg", "Avg kg", "Total CBM", "Avg CBM", "Units",
        "Placement fee", "Currency", "Name"
    ]
    for col_idx, h in enumerate(data_headers, 1):
        cell = ws4.cell(row=1, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        ws4.column_dimensions[get_column_letter(col_idx)].width = 15
    for r_idx, ship in enumerate(df_selected.itertuples(), 2):
        ws4.cell(row=r_idx, column=1,  value=ship.fba_id)
        ws4.cell(row=r_idx, column=2,  value=ship.shipment_id)
        ws4.cell(row=r_idx, column=3,  value=ship.fc)
        ws4.cell(row=r_idx, column=4,  value=ship.status)
        ws4.cell(row=r_idx, column=5,  value=ship.city)
        ws4.cell(row=r_idx, column=6,  value=ship.state)
        ws4.cell(row=r_idx, column=7,  value=ship.postal)
        ws4.cell(row=r_idx, column=8,  value=ship.country)
        ws4.cell(row=r_idx, column=9,  value=ship.line1)
        ws4.cell(row=r_idx, column=10, value=ship.source_city)
        ws4.cell(row=r_idx, column=11, value=ship.source_country)
        ws4.cell(row=r_idx, column=12, value=str(ship.dw_start)[:10] if ship.dw_start else "")
        ws4.cell(row=r_idx, column=13, value=str(ship.dw_end)[:10] if ship.dw_end else "")
        ws4.cell(row=r_idx, column=14, value=int(ship.box_count))
        ws4.cell(row=r_idx, column=15, value=round(ship.total_kg, 2))
        ws4.cell(row=r_idx, column=16, value=round(ship.avg_kg, 2))
        ws4.cell(row=r_idx, column=17, value=round(ship.total_cbm, 4))
        ws4.cell(row=r_idx, column=18, value=round(ship.avg_cbm, 4))
        ws4.cell(row=r_idx, column=19, value=int(ship.total_units))
        ws4.cell(row=r_idx, column=20, value=getattr(ship, "fee_amount", ""))
        ws4.cell(row=r_idx, column=21, value=getattr(ship, "fee_currency", ""))
        ws4.cell(row=r_idx, column=22, value=ship.name)
    ws4.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ============================================
# МОДУЛЬ 2: Парсер квот перевізників
# ============================================

# Маппінг назв листів → код країни
_SHEET_COUNTRY_MAP = {
    "us": "US", "us sea+truck": "US", "us sea + spd": "US",
    "ca": "CA",
    "au": "AU",
    "jp": "JP",
    "uk&de": None,
    "uk&de-tax included": None,
}
_SKIP_SHEETS = {"12kg (the same price as 100kg)", "compensation terms"}

def _parse_delivery(text):
    """'After ETD 18-22 days' | '30-35' | '15-21' → (15, 21)"""
    if not text:
        return None, None
    m = re.search(r'(\d+)\s*[-–]\s*(\d+)', str(text))
    return (int(m.group(1)), int(m.group(2))) if m else (None, None)

def _parse_truck_text_rate(text):
    """ ' BY TRUCK :$1.57  33-40day' → (1.57, 33, 40) """
    m_rate = re.search(r'\$(\d+\.?\d*)', str(text))
    m_days = re.search(r'(\d+)\s*[-–]\s*(\d+)', str(text))
    return (
        float(m_rate.group(1)) if m_rate else None,
        int(m_days.group(1))   if m_days else None,
        int(m_days.group(2))   if m_days else None,
    )

def _parse_vertical_sheet(ws, country, carrier_name, quote_date, file_name):
    rows = []
    current_fc_group = "standard"
    current_service  = None
    has_800 = False

    for row in ws.iter_rows(values_only=True):
        c1 = row[1] if len(row) > 1 else None
        c2 = row[2] if len(row) > 2 else None
        c3 = row[3] if len(row) > 3 else None
        c4 = row[4] if len(row) > 4 else None

        if c1 is None and c2 is None:
            continue
        if isinstance(c1, str) and c1.strip().startswith('1.'):
            continue

        # Рядок-заголовок секції (col2 містить 'kg')
        if isinstance(c2, str) and 'kg' in c2.lower():
            svc_name = str(c1).strip() if c1 else ""
            if any(x in svc_name for x in
                   ['IUSP','IUSQ','IUSJ','IUSF','IUSL','IUST','IUSR','91730','Sumac']):
                current_fc_group = svc_name
                current_service  = None
            else:
                current_fc_group = "standard"
                current_service  = svc_name
            has_800 = isinstance(c3, str) and 'kg' in c3.lower()
            continue

        # Рядок тарифу (col2 = число)
        if isinstance(c2, (int, float)):
            zone = str(c1).strip() if c1 else ""
            if not zone or 'Non Amazon' in zone:
                continue
            rate_101 = float(c2)
            if has_800:
                rate_800 = float(c3) if isinstance(c3, (int, float)) else None
                delivery_raw = c4
            else:
                rate_800 = None
                delivery_raw = c3
            dmin, dmax = _parse_delivery(delivery_raw)
            svc = zone if current_fc_group != "standard" else current_service
            rows.append({
                "carrier_name":      carrier_name,
                "quote_date":        quote_date,
                "marketplace":       country,
                "fc_group":          current_fc_group,
                "service_type":      svc,
                "zone":              zone if current_fc_group == "standard" else None,
                "rate_101kg":        rate_101,
                "rate_800kg":        rate_800,
                "delivery_days_min": dmin,
                "delivery_days_max": dmax,
                "raw_delivery_text": str(delivery_raw) if delivery_raw else None,
                "file_name":         file_name,
            })
            continue

        # Baitong: тариф у текстовому рядку col4 " BY TRUCK :$1.57"
        if (isinstance(c1, str) and c2 is None and
                c4 and isinstance(c4, str) and '$' in c4):
            rate, dmin, dmax = _parse_truck_text_rate(c4)
            if rate:
                rows.append({
                    "carrier_name":      carrier_name,
                    "quote_date":        quote_date,
                    "marketplace":       country,
                    "fc_group":          current_fc_group,
                    "service_type":      str(c1).strip(),
                    "zone":              None,
                    "rate_101kg":        rate,
                    "rate_800kg":        None,
                    "delivery_days_min": dmin,
                    "delivery_days_max": dmax,
                    "raw_delivery_text": c4.strip(),
                    "file_name":         file_name,
                })
    return rows

def _parse_ukde_sheet(ws, carrier_name, quote_date, file_name):
    """Горизонтальний формат UK&DE: Country | AIR | time | SEA | time | TRUCK | time"""
    rows = []
    service_map = {}

    for row in ws.iter_rows(values_only=True):
        c1 = row[1] if len(row) > 1 else None
        if c1 is None:
            continue
        if isinstance(c1, str) and c1.strip() == 'Country':
            for ci, v in enumerate(row[2:], 2):
                if isinstance(v, str) and v.strip() in ('AIR','Regular SEA','TRUCK','SEA'):
                    service_map[ci] = v.strip()
            continue
        if isinstance(c1, str) and 'FBA' in c1:
            country_code = "UK" if c1.startswith('UK') else "DE"
            for ci, svc in service_map.items():
                rate     = row[ci]     if ci < len(row) else None
                time_val = row[ci + 1] if (ci + 1) < len(row) else None
                if isinstance(rate, (int, float)) and rate:
                    dmin, dmax = _parse_delivery(time_val)
                    rows.append({
                        "carrier_name":      carrier_name,
                        "quote_date":        quote_date,
                        "marketplace":       country_code,
                        "fc_group":          "standard",
                        "service_type":      svc,
                        "zone":              "FBA",
                        "rate_101kg":        float(rate),
                        "rate_800kg":        None,
                        "delivery_days_min": dmin,
                        "delivery_days_max": dmax,
                        "raw_delivery_text": str(time_val) if time_val else None,
                        "file_name":         file_name,
                    })
    return rows

def parse_carrier_excel(file_path: str, carrier_name: str,
                        quote_date: datetime.date = None) -> list:
    """
    Парсить Excel-файл перевізника → list[dict] для INSERT в tender_quotes.
    Підтримує: Maximumstores, UnrealChina, Baitong, Quotations2026 (формули).
    """
    if quote_date is None:
        quote_date = datetime.date.today()
    file_name = os.path.basename(file_path)
    wb = load_workbook(file_path, data_only=True)   # data_only=True → формули як значення
    all_rows = []

    for sheet_name in wb.sheetnames:
        sheet_key = sheet_name.strip().lower()
        if sheet_key in _SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        country = _SHEET_COUNTRY_MAP.get(sheet_key)
        if country is None and ('uk' in sheet_key or 'de' in sheet_key):
            all_rows.extend(_parse_ukde_sheet(ws, carrier_name, quote_date, file_name))
        elif country:
            all_rows.extend(_parse_vertical_sheet(ws, country, carrier_name, quote_date, file_name))

    return all_rows

_CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS tender_quotes (
    id                 SERIAL PRIMARY KEY,
    carrier_name       TEXT        NOT NULL,
    quote_date         DATE        NOT NULL,
    marketplace        TEXT        NOT NULL,
    fc_group           TEXT,
    service_type       TEXT,
    zone               TEXT,
    rate_101kg         NUMERIC(10,4),
    rate_800kg         NUMERIC(10,4),
    delivery_days_min  INTEGER,
    delivery_days_max  INTEGER,
    raw_delivery_text  TEXT,
    file_name          TEXT,
    uploaded_at        TIMESTAMP DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_tq_carrier  ON tender_quotes(carrier_name);
CREATE INDEX IF NOT EXISTS idx_tq_market   ON tender_quotes(marketplace);
CREATE INDEX IF NOT EXISTS idx_tq_date     ON tender_quotes(quote_date);
"""

def _ensure_table(conn):
    with conn.cursor() as cur:
        cur.execute(_CREATE_TABLE_SQL)
        conn.commit()

def _load_quotes_to_db(rows: list, conn) -> int:
    if not rows:
        return 0
    carrier = rows[0]["carrier_name"]
    qdate   = rows[0]["quote_date"]
    with conn.cursor() as cur:
        cur.execute(
            "DELETE FROM tender_quotes WHERE carrier_name = %s AND quote_date = %s",
            (carrier, qdate)
        )
        cur.executemany("""
            INSERT INTO tender_quotes
                (carrier_name, quote_date, marketplace, fc_group, service_type,
                 zone, rate_101kg, rate_800kg, delivery_days_min, delivery_days_max,
                 raw_delivery_text, file_name)
            VALUES
                (%(carrier_name)s, %(quote_date)s, %(marketplace)s, %(fc_group)s,
                 %(service_type)s, %(zone)s, %(rate_101kg)s, %(rate_800kg)s,
                 %(delivery_days_min)s, %(delivery_days_max)s,
                 %(raw_delivery_text)s, %(file_name)s)
        """, rows)
        conn.commit()
    return len(rows)


# ============================================
# МОДУЛЬ 2 UI: секція завантаження квот
# ============================================
KNOWN_CARRIERS = [
    "Maximumstores",
    "UnrealChina",
    "Baitong",
    "Інший (ввести вручну)",
]

def _render_quote_upload():
    """Секція '📥 Квоти перевізників' всередині show_tender_tab()"""
    st.markdown("---")
    st.subheader("📥 Квоти перевізників")
    st.caption("Завантаж Excel від перевізника — система сама розпарсить тарифи і збереже в БД.")

    col1, col2, col3 = st.columns([2, 2, 2])
    with col1:
        choice = st.selectbox("Перевізник", KNOWN_CARRIERS, key="carrier_select")
        carrier_name = (
            st.text_input("Назва перевізника", key="carrier_custom")
            if choice == "Інший (ввести вручну)" else choice
        )
    with col2:
        quote_date = st.date_input("Дата квоти", value=datetime.date.today(), key="quote_date")
    with col3:
        uploaded = st.file_uploader("Excel-файл від перевізника", type=["xlsx"], key="quote_file")

    if not (uploaded and carrier_name):
        # ── Поточні квоти в БД ──────────────────────────────────────────────
        _render_quotes_table()
        return

    # Кнопка "Розпарсити і перевірити"
    if st.button("🔍 Розпарсити і перевірити", type="secondary", key="btn_parse"):
        with st.spinner("Парсимо файл..."):
            try:
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                    tmp.write(uploaded.getvalue())
                    tmp_path = tmp.name

                rows = parse_carrier_excel(tmp_path, carrier_name, quote_date)
                os.unlink(tmp_path)

                if not rows:
                    st.error("❌ Не вдалось витягти дані. Перевір формат файлу.")
                    return

                st.session_state["_tq_pending"] = rows
                st.session_state["_tq_carrier"] = carrier_name

                df_prev = pd.DataFrame(rows)
                markets = df_prev["marketplace"].value_counts().to_dict()

                c_a, c_b, c_c = st.columns(3)
                c_a.metric("Рядків тарифів", len(rows))
                c_b.metric("Ринків", len(markets))
                c_c.metric("Перевізник", carrier_name)

                st.dataframe(
                    df_prev[[
                        "marketplace", "fc_group", "service_type", "zone",
                        "rate_101kg", "rate_800kg",
                        "delivery_days_min", "delivery_days_max"
                    ]],
                    use_container_width=True,
                    height=280,
                )
            except Exception as e:
                st.error(f"❌ Помилка парсингу: {e}")

    # Кнопки після preview: скачати і/або зберегти в БД
    pending = st.session_state.get("_tq_pending", [])
    if pending and st.session_state.get("_tq_carrier") == carrier_name:

        # Генеруємо Excel з розпарсених даних
        df_dl = pd.DataFrame(pending)
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df_dl.to_excel(writer, sheet_name="Тарифи", index=False)
        buf.seek(0)

        col_dl, col_sv = st.columns([1, 1])
        with col_dl:
            st.download_button(
                label="📥 Скачати як Excel",
                data=buf.getvalue(),
                file_name=f"{carrier_name}_{quote_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with col_sv:
            if st.button(
                f"✅ Зберегти {len(pending)} тарифів у БД",
                type="primary", key="btn_save_quotes",
                use_container_width=True,
            ):
                try:
                    conn = _get_conn()
                    _ensure_table(conn)
                    n = _load_quotes_to_db(pending, conn)
                    conn.close()
                    st.success(
                        f"✅ Збережено **{n}** рядків від **{carrier_name}** "
                        f"на {quote_date.strftime('%d.%m.%Y')}"
                    )
                    st.session_state.pop("_tq_pending", None)
                    st.session_state.pop("_tq_carrier", None)
                    st.cache_data.clear()
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Помилка БД: {e}")

    _render_quotes_table()


def _render_quotes_table():
    """Зведена таблиця всіх квот що є в БД + кнопка скачати Excel"""
    st.markdown("---")
    st.markdown("#### 📊 Квоти в базі даних")
    try:
        conn = _get_conn()

        # Зведена (summary) таблиця
        df_summary = pd.read_sql("""
            SELECT
                carrier_name        AS "Перевізник",
                quote_date          AS "Дата квоти",
                marketplace         AS "Ринок",
                COUNT(*)            AS "Тарифів",
                ROUND(MIN(rate_101kg)::numeric, 3) AS "Min $/kg",
                ROUND(MAX(rate_101kg)::numeric, 3) AS "Max $/kg",
                MIN(delivery_days_min)              AS "Min днів",
                MAX(delivery_days_max)              AS "Max днів"
            FROM tender_quotes
            GROUP BY carrier_name, quote_date, marketplace
            ORDER BY quote_date DESC, carrier_name, marketplace
        """, conn)

        # Повна таблиця для Excel
        df_full = pd.read_sql("""
            SELECT
                carrier_name        AS "Перевізник",
                quote_date          AS "Дата квоти",
                marketplace         AS "Ринок",
                fc_group            AS "FC Group",
                service_type        AS "Сервіс",
                zone                AS "Зона",
                rate_101kg          AS "$/kg (101kg+)",
                rate_800kg          AS "$/kg (800kg+)",
                delivery_days_min   AS "Днів мін",
                delivery_days_max   AS "Днів макс",
                raw_delivery_text   AS "Доставка (оригінал)",
                uploaded_at         AS "Завантажено"
            FROM tender_quotes
            ORDER BY carrier_name, marketplace, service_type, zone
        """, conn)
        conn.close()

        if df_summary.empty:
            st.info("Квоти ще не завантажені. Завантаж перший файл вище.")
            return

        # Показуємо зведену таблицю
        st.dataframe(df_summary, use_container_width=True, height=280)

        # Кнопка скачати повну матрицю в Excel
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            # Sheet 1: Матриця порівняння (pivot: перевізники vs сервіси)
            df_pivot = df_full.copy()
            df_pivot.to_excel(writer, sheet_name="Всі тарифи", index=False)

            # Sheet 2: Зведена по перевізникам
            df_summary.to_excel(writer, sheet_name="Зведена", index=False)

            # Sheet 3: US порівняння — найкорисніше для вибору
            df_us = df_full[df_full["Ринок"] == "US"][[
                "Перевізник", "Сервіс", "Зона",
                "$/kg (101kg+)", "$/kg (800kg+)",
                "Днів мін", "Днів макс"
            ]].sort_values(["$/kg (101kg+)", "Днів мін"])
            if not df_us.empty:
                df_us.to_excel(writer, sheet_name="US порівняння", index=False)

        buf.seek(0)
        fname = f"tender_quotes_{datetime.date.today().isoformat()}.xlsx"
        st.download_button(
            label="📥 Скачати матрицю квот (Excel)",
            data=buf.getvalue(),
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    except Exception:
        st.info("Таблиця tender_quotes буде створена при першому завантаженні квоти.")


# ============================================
# Streamlit UI — головна функція
# ============================================
def _render_ai_analysis():
    """Суб-таб: AI аналіз квот — 3 варіанти прямо в Tender Tab"""
    st.markdown("### 🤖 AI аналіз квот перевізників")
    st.caption("Агент читає актуальні квоти з БД і шипменти — видає 3 варіанти рішення.")

    if st.button("🔍 Запустити AI аналіз", type="primary", use_container_width=True):
        with st.spinner("Аналізуємо квоти та шипменти..."):
            try:
                conn = _get_conn()

                # Квоти з БД (остання дата)
                df_quotes = pd.read_sql("""
                    SELECT carrier_name, marketplace, service_type, zone,
                           rate_101kg, rate_800kg,
                           delivery_days_min, delivery_days_max
                    FROM tender_quotes
                    WHERE quote_date = (SELECT MAX(quote_date) FROM tender_quotes)
                    ORDER BY marketplace, rate_101kg
                """, conn)

                # Активні шипменти
                df_ships = pd.read_sql("""
                    SELECT destination_fc AS fc, ship_to_country AS country,
                           COUNT(DISTINCT b.box_id) AS boxes,
                           SUM(NULLIF(b.weight_kg,'')::float) AS total_kg,
                           SUM(NULLIF(b.volume_cbm,'')::float) AS total_cbm
                    FROM public.fba_inbound_shipments_v2 s
                    LEFT JOIN public.fba_shipment_boxes b ON b.inbound_plan_id = s.inbound_plan_id
                    WHERE s.status IN ('WORKING','READY_TO_SHIP','RECEIVING')
                    GROUP BY s.destination_fc, s.ship_to_country
                    HAVING COUNT(DISTINCT b.box_id) > 0
                """, conn)
                conn.close()

                if df_quotes.empty:
                    st.warning("⚠️ Квоти не завантажені. Спочатку завантаж файли від перевізників у суб-табі 'Квоти перевізників'.")
                    return

                total_kg  = df_ships["total_kg"].sum() if not df_ships.empty else 0
                total_cbm = df_ships["total_cbm"].sum() if not df_ships.empty else 0
                total_boxes = int(df_ships["boxes"].sum()) if not df_ships.empty else 0

                quotes_text   = df_quotes.to_string(index=False)
                shipments_text = df_ships.to_string(index=False) if not df_ships.empty else "Немає активних шипментів"

                prompt = f"""Ти — AI агент з логістики. Проаналізуй квоти перевізників і активні шипменти Amazon FBA.

АКТИВНІ ШИПМЕНТИ:
{shipments_text}

Загальний вантаж: {total_boxes} коробок | {total_kg:.0f} kg | {total_cbm:.2f} CBM

АКТУАЛЬНІ КВОТИ ПЕРЕВІЗНИКІВ ($/kg):
{quotes_text}

Надай рівно 3 варіанти у форматі нижче. Для кожного варіанту порахуй орієнтовну вартість ({total_kg:.0f} kg × ставка).

ВАРІАНТ A 💰 МІНІМАЛЬНА ЦІНА
Перевізник: [назва]
Сервіс: [тип]
Ставка: $[X]/kg
Орієнтовна вартість: $[сума]
Термін доставки: [X-Y] днів
Чому: [1 речення]

ВАРІАНТ B ⚡ МАКСИМАЛЬНА ШВИДКІСТЬ
Перевізник: [назва]
Сервіс: [тип]
Ставка: $[X]/kg
Орієнтовна вартість: $[сума]
Термін доставки: [X-Y] днів
Чому: [1 речення]

ВАРІАНТ C 🔀 ДИВЕРСИФІКАЦІЯ
Перевізник 1: [назва] — [X]% вантажу
Перевізник 2: [назва] — [Y]% вантажу
Орієнтовна вартість: $[сума]
Термін доставки: [X-Y] днів
Чому: [1 речення]

РЕКОМЕНДАЦІЯ: [A/B/C] — [одне речення чому]"""

                # Виклик Claude API
                import requests as req
                try:
                    _akey = st.secrets["ANTHROPIC_API_KEY"]
                except Exception:
                    _akey = os.getenv("ANTHROPIC_API_KEY", "")
                resp = req.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={
                        "Content-Type": "application/json",
                        "x-api-key": _akey,
                        "anthropic-version": "2023-06-01"
                    },
                    json={
                        "model": "claude-sonnet-4-6",
                        "max_tokens": 1500,
                        "messages": [{"role": "user", "content": prompt}]
                    },
                    timeout=60
                )
                result = resp.json()
                if "error" in result:
                    raise Exception(result["error"].get("message", str(result["error"])))
                ai_text = result["content"][0]["text"]

                st.session_state["_tender_ai_result"] = ai_text

            except Exception as e:
                st.error(f"❌ Помилка: {e}")

    # Показуємо результат
    if "ـtender_ai_result" in st.session_state or "_tender_ai_result" in st.session_state:
        ai_text = st.session_state.get("_tender_ai_result", "")
        if ai_text:
            # Парсимо 3 варіанти і показуємо в колонках
            col_a, col_b, col_c = st.columns(3)

            def extract_block(text, header):
                lines = text.split('\n')
                result, capture = [], False
                for line in lines:
                    if header in line:
                        capture = True
                    elif capture and line.startswith('ВАРІАНТ') and header not in line:
                        break
                    elif capture and line.startswith('РЕКОМЕНДАЦІЯ'):
                        break
                    elif capture:
                        result.append(line)
                return '\n'.join(result).strip()

            block_a = extract_block(ai_text, "ВАРІАНТ A")
            block_b = extract_block(ai_text, "ВАРІАНТ B")
            block_c = extract_block(ai_text, "ВАРІАНТ C")
            rec = next((l for l in ai_text.split('\n') if 'РЕКОМЕНДАЦІЯ' in l), "")

            with col_a:
                st.markdown("### 💰 Мінімальна ціна")
                st.info(block_a or "—")
            with col_b:
                st.markdown("### ⚡ Швидкість")
                st.warning(block_b or "—")
            with col_c:
                st.markdown("### 🔀 Диверсифікація")
                st.success(block_c or "—")

            if rec:
                st.markdown(f"---\n**{rec}**")


def show_tender_tab():
    st.subheader("📋 Логістичний тендер")

    sub1, sub2, sub3 = st.tabs(["🚢 Тендер", "📥 Квоти перевізників", "🤖 AI Аналіз"])

    with sub2:
        _render_quote_upload()

    with sub3:
        _render_ai_analysis()

    with sub1:
        st.caption(
            "Генерація Excel-файлу для перевізників. "
            "Дані автоматично з БД — формат 1:1 як у попередніх тендерах."
        )

    # --- Fetch data ---
        try:
            df = fetch_tender_shipments()
            fees_df = fetch_placement_fees()
        except Exception as e:
            st.error(f"❌ Не вдалося завантажити дані: {e}")
            st.info("Запусти `14_fba_inbound_v2024_loader.py` щоб наповнити таблиці.")
            return

        if df.empty:
            st.warning("Жодного shipment у `fba_inbound_shipments_v2`. Запусти v2024 loader.")
            return

        df = df.merge(fees_df, on="placement_option_id", how="left")
        df["fee_amount"]   = df["fee_amount"].fillna("")
        df["fee_currency"] = df["fee_currency"].fillna("")

        # --- Filters ---
        col_f1, col_f2, col_f3, col_f4 = st.columns([2, 2, 2, 1])
        with col_f1:
            fc_options = sorted(df["fc"].dropna().unique())
            selected_fcs = st.multiselect("🏠 Destination FC", options=fc_options, default=fc_options)
        with col_f2:
            status_options = sorted(df["status"].dropna().unique())
            default_statuses = [s for s in status_options
                                if s in ("READY_TO_SHIP", "WORKING", "RECEIVING")]
            selected_statuses = st.multiselect(
                "📦 Status", options=status_options,
                default=default_statuses or status_options
            )
        with col_f3:
            marketplace_options = sorted(df["marketplace"].dropna().unique())
            selected_markets = st.multiselect(
                "🌍 Marketplace", options=marketplace_options, default=marketplace_options
            )
        with col_f4:
            st.write("")
            st.write("")
            if st.button("🔄 Оновити"):
                st.cache_data.clear()
                st.rerun()

        mask = (
            df["fc"].isin(selected_fcs) &
            df["status"].isin(selected_statuses) &
            df["marketplace"].isin(selected_markets)
        )
        df_filtered = df[mask].copy()

        if df_filtered.empty:
            st.warning("Під фільтри нічого не підходить. Розкрий фільтри ширше.")
            return

        # --- Metrics ---
        st.markdown("---")
        col_m1, col_m2, col_m3, col_m4, col_m5 = st.columns(5)
        col_m1.metric("🚚 Shipments", len(df_filtered))
        col_m2.metric("📦 Boxes",     int(df_filtered["box_count"].sum()))
        col_m3.metric("⚖️ Weight (kg)",  f"{df_filtered['total_kg'].sum():,.1f}")
        col_m4.metric("📐 Volume (CBM)", f"{df_filtered['total_cbm'].sum():,.2f}")
        col_m5.metric("🔢 Units",        int(df_filtered["total_units"].sum()))

        # --- Shipment selection ---
        st.markdown("### Оберіть shipments для тендера")
        df_display = df_filtered[[
            "fba_id", "fc", "status", "city", "state", "country",
            "dw_start", "dw_end", "box_count", "total_kg", "total_cbm",
            "total_units", "fee_amount"
        ]].copy()
        df_display.insert(0, "✓", True)
        df_display["dw_start"]   = df_display["dw_start"].astype(str).str[:10]
        df_display["dw_end"]     = df_display["dw_end"].astype(str).str[:10]
        df_display["total_kg"]   = df_display["total_kg"].round(1)
        df_display["total_cbm"]  = df_display["total_cbm"].round(3)

        edited = st.data_editor(
            df_display,
            hide_index=True,
            use_container_width=True,
            column_config={
                "✓":          st.column_config.CheckboxColumn(width="small"),
                "fba_id":     st.column_config.TextColumn("FBA ID",   width="medium"),
                "fc":         st.column_config.TextColumn("FC",       width="small"),
                "status":     st.column_config.TextColumn("Status"),
                "city":       st.column_config.TextColumn("City"),
                "state":      st.column_config.TextColumn("St"),
                "country":    st.column_config.TextColumn("Cty"),
                "dw_start":   st.column_config.TextColumn("DW start"),
                "dw_end":     st.column_config.TextColumn("DW end"),
                "box_count":  st.column_config.NumberColumn("Boxes"),
                "total_kg":   st.column_config.NumberColumn("kg",  format="%.1f"),
                "total_cbm":  st.column_config.NumberColumn("CBM", format="%.3f"),
                "total_units":st.column_config.NumberColumn("Units"),
                "fee_amount": st.column_config.TextColumn("Fee"),
            },
            disabled=["fba_id","fc","status","city","state","country",
                      "dw_start","dw_end","box_count","total_kg","total_cbm",
                      "total_units","fee_amount"],
            key="tender_editor",
        )

        selected_fba_ids = edited[edited["✓"]]["fba_id"].tolist()
        df_selected = df_filtered[df_filtered["fba_id"].isin(selected_fba_ids)].copy()

        if df_selected.empty:
            st.warning("⚠️ Жодного shipment не вибрано. Постав галочку хоча б на одному.")
            return

        # --- Excel generation ---
        st.markdown("---")
        st.markdown("### 📥 Згенерувати Excel")
        col_g1, col_g2, col_g3 = st.columns([2, 2, 2])
        with col_g1:
            pickup_date_obj = st.date_input(
                "📅 Pick-up date",
                value=datetime.date.today() + datetime.timedelta(days=7),
            )
            pickup_date_str = pickup_date_obj.strftime("%d.%m.%y")
        with col_g2:
            output_name = st.text_input(
                "📄 Filename",
                value=f"tender_{datetime.date.today().isoformat()}.xlsx",
            )
        with col_g3:
            st.write("")
            st.write("")
            st.info(f"✅ Відібрано **{len(df_selected)}** shipments")

        excel_bytes = build_tender_excel(df_selected, pickup_date_str)
        st.download_button(
            label=f"📥 Завантажити {output_name}",
            data=excel_bytes,
            file_name=output_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )

        with st.expander("👁️ Preview — що буде в Excel (перші 3 рядки)"):
            for _, ship in df_selected.head(3).iterrows():
                st.markdown(
                    f"**{ship['fba_id']}** · {ship['fc']} · "
                    f"{ship['box_count']} boxes · {int(ship['total_units'])} units"
                )
                parts = [ship["fc"]]
                if ship.get("line1"):
                    parts.append(f"{ship['line1']} {ship.get('postal','')}".strip())
                parts.append(f"{(ship.get('city') or '').upper()}, {ship.get('state','')}")
                if ship.get("country"):
                    parts.append("United States" if ship["country"] == "US" else ship["country"])
                st.caption(f"Ship to: {' - '.join(parts)}")
                dw = _format_dw(ship.get("dw_start"), ship.get("dw_end"))
                if dw:
                    st.caption(f"Delivery window: {dw}")
                st.markdown("---")
