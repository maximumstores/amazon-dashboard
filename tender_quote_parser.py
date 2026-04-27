"""
tender_quote_parser.py
Универсальный парсер квот перевозчиков → список dict для записи в PostgreSQL

Поддерживает:
  - Maximumstores (US/CA/AU/UK/JP, 1 вес, "After ETD 18-22 days")
  - UnrealChina   (US/CA/UK, 2 веса 101/800kg, "30-35")
  - Baitong       (US only, 1 вес, BY TRUCK :$1.57 в тексте)
  - Quotations2026 (US/CA/UK, 2 веса, Excel-формулы через data_only)
"""

import re
from datetime import date
from openpyxl import load_workbook

# ─── Маппинг имён листов → код страны ───────────────────────────────────────
SHEET_COUNTRY_MAP = {
    "us":  "US", "us sea+truck": "US", "us sea + spd": "US",
    "ca":  "CA",
    "au":  "AU",
    "uk&de": None,       # горизонтальный формат, разбирается отдельно
    "uk&de-tax included": None,
    "jp":  "JP",
}

SKIP_SHEETS = {"12kg (the same price as 100kg)", "compensation terms"}

# ─── Парсинг строки доставки → (min, max) ───────────────────────────────────
def parse_delivery(text):
    """'After ETD 18-22 days' | '30-35' | '15-21' → (15, 21)"""
    if not text:
        return None, None
    m = re.search(r'(\d+)\s*[-–]\s*(\d+)', str(text))
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None

# ─── Парсинг тарифа из текстовой строки Baitong ─────────────────────────────
def parse_truck_text_rate(text):
    """' BY TRUCK :$1.57  33-40day' → (1.57, 33, 40)"""
    rate, dmin, dmax = None, None, None
    m_rate = re.search(r'\$(\d+\.?\d*)', str(text))
    m_days = re.search(r'(\d+)\s*[-–]\s*(\d+)', str(text))
    if m_rate:
        rate = float(m_rate.group(1))
    if m_days:
        dmin, dmax = int(m_days.group(1)), int(m_days.group(2))
    return rate, dmin, dmax

# ─── Вертикальный формат (US / CA / AU / JP) ────────────────────────────────
def parse_vertical_sheet(ws, country, carrier_name, quote_date, file_name):
    """
    Паттерн строки-заголовка секции:
        col1=service_name, col2='101kg+' [, col3='800kg+'], col_last='Delivery time'
    Паттерн строки тарифа:
        col1=zone, col2=rate_101 (float) [, col3=rate_800], col_last=delivery_str
    """
    rows = []
    current_fc_group  = "standard"
    current_service   = None
    has_800           = False

    for row in ws.iter_rows(values_only=True):
        c1, c2, c3 = row[1], row[2], row[3] if len(row) > 3 else None
        c4 = row[4] if len(row) > 4 else None

        # пропускаем полностью пустые строки и строки-комментарии
        if c1 is None and c2 is None:
            continue
        if isinstance(c1, str) and c1.strip().startswith('1.'):
            continue

        # ── строка-заголовок секции ──────────────────────────────────────────
        if isinstance(c2, str) and 'kg' in c2.lower():
            svc_name = str(c1).strip() if c1 else ""
            # Определяем: это FC-группа (IUSP/IUSQ/...) или сервис верхнего уровня
            if any(x in svc_name for x in ['IUSP', 'IUSQ', 'IUSJ', 'IUSF', 'IUSL',
                                             'IUST', 'IUSR', '91730', 'Sumac']):
                current_fc_group = svc_name
                current_service  = None   # будет уточнён в строках тарифа
            else:
                current_fc_group = "standard"
                current_service  = svc_name
            has_800 = isinstance(c3, str) and 'kg' in c3.lower()
            continue

        # ── строка тарифа (col2 = число) ────────────────────────────────────
        if isinstance(c2, (int, float)):
            zone = str(c1).strip() if c1 else ""
            if not zone or 'Non Amazon' in zone:
                continue

            rate_101 = float(c2) if isinstance(c2, (int, float)) else None

            if has_800:
                rate_800 = float(c3) if isinstance(c3, (int, float)) else None
                delivery_raw = c4
            else:
                rate_800 = None
                delivery_raw = c3

            dmin, dmax = parse_delivery(delivery_raw)

            # При FC-группе col1 — имя сервиса (Matson CLX, EXX/ZIM...)
            svc = zone if current_fc_group != "standard" else current_service

            rows.append({
                "carrier_name":       carrier_name,
                "quote_date":         quote_date,
                "marketplace":        country,
                "fc_group":           current_fc_group,
                "service_type":       svc,
                "zone":               zone if current_fc_group == "standard" else None,
                "rate_101kg":         rate_101,
                "rate_800kg":         rate_800,
                "delivery_days_min":  dmin,
                "delivery_days_max":  dmax,
                "raw_delivery_text":  str(delivery_raw) if delivery_raw else None,
                "file_name":          file_name,
            })
            continue

        # ── Baitong: тариф зашит в текст col4 (" BY TRUCK :$1.57") ─────────
        if isinstance(c1, str) and c2 is None and c4 and isinstance(c4, str) and '$' in c4:
            rate, dmin, dmax = parse_truck_text_rate(c4)
            if rate:
                rows.append({
                    "carrier_name":       carrier_name,
                    "quote_date":         quote_date,
                    "marketplace":        country,
                    "fc_group":           current_fc_group,
                    "service_type":       str(c1).strip(),
                    "zone":               None,
                    "rate_101kg":         rate,
                    "rate_800kg":         None,
                    "delivery_days_min":  dmin,
                    "delivery_days_max":  dmax,
                    "raw_delivery_text":  c4.strip(),
                    "file_name":          file_name,
                })

    return rows

# ─── Горизонтальный формат UK&DE ─────────────────────────────────────────────
def parse_ukde_sheet(ws, carrier_name, quote_date, file_name):
    """
    Формат: Country | AIR | time | SEA | time | TRUCK | time
    Строки: UK-FBA, DE-FBA и т.д.
    """
    rows = []
    service_map = {}  # col_index → service_type

    for row in ws.iter_rows(values_only=True):
        c1 = row[1]
        if c1 is None:
            continue

        # строка заголовков сервисов
        if isinstance(c1, str) and c1.strip() == 'Country':
            for ci, v in enumerate(row[2:], 2):
                if isinstance(v, str) and v.strip() in ('AIR', 'Regular SEA', 'TRUCK', 'SEA'):
                    service_map[ci] = v.strip()
            continue

        # строка тарифа (UK-FBA, DE-FBA)
        if isinstance(c1, str) and 'FBA' in c1:
            country_code = "UK" if c1.startswith('UK') else "DE"
            for ci, svc in service_map.items():
                rate = row[ci] if ci < len(row) else None
                time_val = row[ci+1] if (ci+1) < len(row) else None
                if isinstance(rate, (int, float)) and rate:
                    dmin, dmax = parse_delivery(time_val)
                    rows.append({
                        "carrier_name":       carrier_name,
                        "quote_date":         quote_date,
                        "marketplace":        country_code,
                        "fc_group":           "standard",
                        "service_type":       svc,
                        "zone":               "FBA",
                        "rate_101kg":         float(rate),
                        "rate_800kg":         None,
                        "delivery_days_min":  dmin,
                        "delivery_days_max":  dmax,
                        "raw_delivery_text":  str(time_val) if time_val else None,
                        "file_name":          file_name,
                    })
    return rows

# ─── Главная функция ──────────────────────────────────────────────────────────
def parse_carrier_excel(file_path: str, carrier_name: str,
                         quote_date: date = None) -> list[dict]:
    """
    Парсит файл перевозчика и возвращает список dict готовых к INSERT в БД.
    
    Args:
        file_path:    путь к .xlsx файлу
        carrier_name: имя перевозчика (выбирает менеджер при загрузке)
        quote_date:   дата квоты (если None — используется сегодня)
    
    Returns:
        list of dicts соответствующих колонкам таблицы tender_quotes
    """
    if quote_date is None:
        quote_date = date.today()
    file_name = file_path.split("/")[-1]

    # Quotations2026 содержит формулы — нужен data_only
    wb = load_workbook(file_path, data_only=True)
    all_rows = []

    for sheet_name in wb.sheetnames:
        sheet_key = sheet_name.strip().lower()

        if sheet_key in SKIP_SHEETS:
            continue

        ws = wb[sheet_name]
        country = SHEET_COUNTRY_MAP.get(sheet_key)

        if country is None and ('uk' in sheet_key or 'de' in sheet_key):
            # горизонтальный UK&DE лист
            all_rows.extend(parse_ukde_sheet(ws, carrier_name, quote_date, file_name))
        elif country:
            all_rows.extend(parse_vertical_sheet(ws, country, carrier_name, quote_date, file_name))
        # остальные листы (JP, AU, 12KG…) — пропускаем или добавим позже

    return all_rows


# ─── Тест на всех 4 файлах ───────────────────────────────────────────────────
if __name__ == "__main__":
    test_files = [
        ("/mnt/user-data/uploads/Quotation_for_Maximumstores_22-Apr.xlsx",  "Maximumstores"),
        ("/mnt/user-data/uploads/UnrealChina_03_20_26_xlsx1.xlsx",          "UnrealChina"),
        ("/mnt/user-data/uploads/Baitong_04_09_26.xlsx",                    "Baitong"),
        ("/mnt/user-data/uploads/Quotations_2026_4_22_.xlsx",               "Quotations2026"),
    ]

    from datetime import date
    grand_total = 0
    for path, name in test_files:
        rows = parse_carrier_excel(path, name, date(2026, 4, 22))
        us_rows = [r for r in rows if r["marketplace"] == "US"]
        ca_rows = [r for r in rows if r["marketplace"] == "CA"]
        uk_rows = [r for r in rows if r["marketplace"] == "UK"]
        de_rows = [r for r in rows if r["marketplace"] == "DE"]
        # строки без rate
        no_rate = [r for r in rows if r["rate_101kg"] is None]
        grand_total += len(rows)

        print(f"\n{'─'*55}")
        print(f"Carrier: {name}  → {len(rows)} строк тарифов")
        print(f"  US={len(us_rows)}  CA={len(ca_rows)}  UK={len(uk_rows)}  DE={len(de_rows)}")
        print(f"  Строк без rate_101kg: {len(no_rate)}")
        if us_rows:
            r = us_rows[0]
            print(f"  Sample US: svc={r['service_type']!r}, zone={r['zone']!r}, "
                  f"rate={r['rate_101kg']}, d={r['delivery_days_min']}-{r['delivery_days_max']}")

    print(f"\n{'='*55}")
    print(f"ИТОГО строк во всех 4 файлах: {grand_total}")


# ═══════════════════════════════════════════════════════════════════════════════
# SQL: CREATE TABLE
# ═══════════════════════════════════════════════════════════════════════════════
CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS tender_quotes (
    id                 SERIAL PRIMARY KEY,
    carrier_name       TEXT        NOT NULL,
    quote_date         DATE        NOT NULL,
    marketplace        TEXT        NOT NULL,   -- US / CA / AU / UK / DE / JP
    fc_group           TEXT,                   -- standard / IUSP/IUSQ/IUSJ / IUSF ...
    service_type       TEXT,                   -- Matson CLX / EXX/ZIM / Regular SEA ...
    zone               TEXT,                   -- Zip Code (8-9) / Vancouver / FBA ...
    rate_101kg         NUMERIC(10,4),          -- USD/kg для 101kg+
    rate_800kg         NUMERIC(10,4),          -- USD/kg для 800kg+ (если есть)
    delivery_days_min  INTEGER,
    delivery_days_max  INTEGER,
    raw_delivery_text  TEXT,
    file_name          TEXT,
    uploaded_at        TIMESTAMP DEFAULT NOW()
);

CREATE INDEX IF NOT EXISTS idx_tender_quotes_carrier   ON tender_quotes(carrier_name);
CREATE INDEX IF NOT EXISTS idx_tender_quotes_market    ON tender_quotes(marketplace);
CREATE INDEX IF NOT EXISTS idx_tender_quotes_date      ON tender_quotes(quote_date);
CREATE INDEX IF NOT EXISTS idx_tender_quotes_service   ON tender_quotes(service_type);
"""

# ═══════════════════════════════════════════════════════════════════════════════
# DB LOADER: записывает строки в PostgreSQL
# ═══════════════════════════════════════════════════════════════════════════════
def load_quotes_to_db(rows: list[dict], conn) -> int:
    """
    Вставляет строки тарифов в таблицу tender_quotes.
    Перед вставкой удаляет старые квоты того же перевозчика на ту же дату
    (upsert по carrier_name + quote_date — всегда актуальные данные).
    
    Returns: количество вставленных строк
    """
    if not rows:
        return 0

    carrier  = rows[0]["carrier_name"]
    qdate    = rows[0]["quote_date"]

    with conn.cursor() as cur:
        # Удаляем старую версию квоты этого перевозчика на эту дату
        cur.execute(
            "DELETE FROM tender_quotes WHERE carrier_name = %s AND quote_date = %s",
            (carrier, qdate)
        )

        insert_sql = """
            INSERT INTO tender_quotes
                (carrier_name, quote_date, marketplace, fc_group, service_type,
                 zone, rate_101kg, rate_800kg, delivery_days_min, delivery_days_max,
                 raw_delivery_text, file_name)
            VALUES
                (%(carrier_name)s, %(quote_date)s, %(marketplace)s, %(fc_group)s,
                 %(service_type)s, %(zone)s, %(rate_101kg)s, %(rate_800kg)s,
                 %(delivery_days_min)s, %(delivery_days_max)s,
                 %(raw_delivery_text)s, %(file_name)s)
        """
        cur.executemany(insert_sql, rows)
        conn.commit()
        return len(rows)
